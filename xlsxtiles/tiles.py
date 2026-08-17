from __future__ import annotations

import base64
import hashlib
import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pymupdf
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from .profile import Sheet, profile_workbook
from .soffice import seed_profile, to_pdf

__all__ = ["Tile", "plan_tiles", "render_workbook", "PNG_DATA_URI_PREFIX"]

# Le vectorizer redimensionne l'image vers un budget de patches ; le texte doit
# rester lisible APRÈS ce redimensionnement, sinon la tuile est du bruit.
PATCH_BUDGET = 1024              # ColQwen : ~1024 patches de 28x28 px
PX_PER_PATCH = 28 * 28
MIN_TEXT_PX = 10.0               # hauteur de capitale minimale après resize
CAP_PX = 11.0                    # hauteur de capitale de Calibri 11 à 96 dpi
MIN_ROWS_PER_TILE = 5

DPI = 150
TRIM_PAD_PT = 2.0                # marge laissée autour de l'encre
PAGE_SAFETY = 1.15               # cf. docstring du module
MARGIN_IN = 0.08

PNG_DATA_URI_PREFIX = "data:image/png;base64,"


@dataclass
class Tile:
    """Une plage de la feuille destinée à devenir une image."""
    index: int
    row_start: int
    row_end: int
    col_start: int
    col_end: int

    @property
    def range_a1(self) -> str:
        return (f"{get_column_letter(self.col_start)}{self.row_start}"
                f":{get_column_letter(self.col_end)}{self.row_end}")


# --------------------------------------------------------------------------
# Plan de découpe
# --------------------------------------------------------------------------

def _budget_px(sheet: Sheet, c0: int, c1: int) -> tuple[float, float]:
    """-> (hauteur max d'une tuile, largeur max d'une bande), en px de FEUILLE.

    On part de la contrainte du vectorizer et on remonte : l'image est réduite
    jusqu'au budget de patches, et cette réduction ne doit pas faire tomber la
    hauteur de capitale sous le seuil de lisibilité. La réduction agissant sur
    les deux axes, la contrainte porte sur une surface.

    Le DPI de rendu se simplifie dans l'équation : rasteriser plus fin oblige à
    réduire d'autant plus, la lisibilité finale est inchangée. On raisonne donc
    dans l'unité de la feuille (96 dpi), la même que `Sheet.width_px` — les
    mélanger décalerait le seuil d'un facteur 1,5.
    """
    max_area = PATCH_BUDGET * PX_PER_PATCH * (CAP_PX / MIN_TEXT_PX) ** 2
    width = sheet.width_px(c0, c1)
    max_height = max_area / width if width > 0 else max_area
    return max_height, max_area ** 0.5


def _row_spans(sheet: Sheet, max_height: float) -> list[tuple[int, int]]:
    """Découpe verticale, en accumulant les hauteurs RÉELLES ligne à ligne.

    Un pas fixe serait faux dès qu'une feuille mélange des hauteurs de ligne —
    cas courant des en-têtes hauts et des lignes de commentaire.
    """
    body_start = sheet.min_row + sheet.header_rows
    if body_start > sheet.max_row:
        return [(sheet.min_row, sheet.max_row)]

    # Sauts de page manuels : l'auteur a déjà dit où couper, on le respecte.
    breaks = [b for b in sheet.row_breaks if body_start <= b < sheet.max_row]
    if breaks:
        spans, prev = [], body_start
        for b in breaks:
            spans.append((prev, b))
            prev = b + 1
        spans.append((prev, sheet.max_row))
        return spans

    budget = max_height - sheet.height_px(sheet.min_row, body_start - 1)
    if budget <= 0:
        budget = max_height

    spans: list[tuple[int, int]] = []
    start, acc, n = body_start, 0.0, 0
    for r in range(body_start, sheet.max_row + 1):
        h = sheet.row_px(r)
        if acc + h > budget and n >= MIN_ROWS_PER_TILE:
            spans.append((start, r - 1))
            start, acc, n = r, 0.0, 0
        acc += h
        n += 1
    if start <= sheet.max_row:
        spans.append((start, sheet.max_row))
    return spans or [(body_start, sheet.max_row)]


def _snap_to_anchors(spans: list[tuple[int, int]],
                     anchors: list[tuple[int, int]]) -> list[tuple[int, int]]:
    """Repousse toute coupe qui tomberait au milieu d'un graphique ou d'une image.

    On préfère une tuile plus haute que le budget à une tuile qui coupe un
    visuel en deux : deux demi-graphiques ne portent aucune information.
    """
    if not anchors or len(spans) < 2:
        return spans
    last_row = spans[-1][1]
    out: list[tuple[int, int]] = []
    start = spans[0][0]
    for i, (_, end) in enumerate(spans):
        if i == len(spans) - 1:
            break
        cut = end
        moved = True
        while moved:                     # une ancre peut en chevaucher une autre
            moved = False
            for a0, a1 in anchors:
                if a0 <= cut < a1:
                    cut, moved = a1, True
        if cut >= last_row:
            break                        # l'ancre absorbe tout le reste
        out.append((start, cut))
        start = cut + 1
    if start <= last_row:
        out.append((start, last_row))
    return out


def _col_spans(sheet: Sheet, max_width: float) -> list[tuple[int, int]]:
    """Découpe horizontale, seulement si la feuille est trop large."""
    c0, c1 = sheet.min_col, sheet.max_col
    if sheet.width_px(c0, c1) <= max_width * 1.6:
        return [(c0, c1)]

    key_w = sheet.width_px(c0, c0 + sheet.freeze_cols - 1) if sheet.freeze_cols else 0.0
    spans, start, acc = [], c0, 0.0
    for c in range(c0, c1 + 1):
        w = sheet.col_px(c)
        if acc + w + key_w > max_width and c > start:
            spans.append((start, c - 1))
            start, acc = c, 0.0
        acc += w
    spans.append((start, c1))
    return spans


def plan_tiles(sheet: Sheet) -> list[Tile]:
    """Plage de chaque tuile, avant tout rendu."""
    max_h, max_w = _budget_px(sheet, sheet.min_col, sheet.max_col)
    rows = _snap_to_anchors(_row_spans(sheet, max_h), sheet.anchors)
    cols = _col_spans(sheet, max_w)
    return [Tile(i, r0, r1, c0, c1)
            for i, ((c0, c1), (r0, r1)) in enumerate(
                (c, r) for c in cols for r in rows)]


# --------------------------------------------------------------------------
# Mise en scène
# --------------------------------------------------------------------------

def _stage(src: Path, sheet: Sheet, band: list[Tile], dest: Path,
           extra_width_px: float) -> None:
    """Écrit un classeur ne contenant que la feuille voulue, paginée sur la bande.

    Une seule conversion soffice pour les N tuiles d'une même bande de colonnes :
    les sauts de page manuels rendent la correspondance page -> tuile
    déterministe. Appeler soffice par tuile coûterait ~2 s x N.
    """
    wb = openpyxl.load_workbook(src)
    for name in list(wb.sheetnames):
        if name != sheet.name:
            del wb[name]
    ws = wb[sheet.name]

    r0 = min(t.row_start for t in band)
    r1 = max(t.row_end for t in band)
    c0 = min(t.col_start for t in band)
    c1 = max(t.col_end for t in band)

    for col, width in sheet.col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.print_area = f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}"
    if sheet.header_rows:
        ws.print_title_rows = f"{sheet.min_row}:{sheet.min_row + sheet.header_rows - 1}"
    if sheet.freeze_cols and c0 > sheet.min_col:
        ws.print_title_cols = (f"{get_column_letter(sheet.min_col)}:"
                               f"{get_column_letter(sheet.min_col + sheet.freeze_cols - 1)}")

    ws.page_setup.scale = 100          # PAS fitToPage : il casse la géométrie
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.pageOrder = "overThenDown"
    ws.print_options.gridLines = True
    ws.print_options.headings = False
    ws.page_margins.left = ws.page_margins.right = MARGIN_IN
    ws.page_margins.top = ws.page_margins.bottom = MARGIN_IN
    ws.page_margins.header = ws.page_margins.footer = 0.0

    # Page dimensionnée sur la plus grande tuile de la bande. Sans ça,
    # LibreOffice garde le format par défaut (A4), repagine selon sa propre
    # logique, et la correspondance page -> tuile saute.
    w_in = (sheet.width_px(c0, c1) + extra_width_px) / 96.0
    h_in = max(sheet.height_px(t.row_start, t.row_end) for t in band) / 96.0
    if sheet.header_rows:
        h_in += sheet.height_px(sheet.min_row,
                                sheet.min_row + sheet.header_rows - 1) / 96.0
    ws.page_setup.paperWidth = f"{(w_in + 2 * MARGIN_IN) * 25.4 * PAGE_SAFETY:.2f}mm"
    ws.page_setup.paperHeight = f"{(h_in + 2 * MARGIN_IN) * 25.4 * PAGE_SAFETY:.2f}mm"
    ws.page_setup.paperSize = None

    # openpyxl : row_breaks est un objet RowBreak, la liste est dans .brk
    ws.row_breaks.brk = []
    ws.col_breaks.brk = []
    for b in sorted({t.row_end for t in band if t.row_end < r1}):
        ws.row_breaks.append(Break(id=b))

    wb.save(dest)


def _clip(page: pymupdf.Page) -> pymupdf.Rect | None:
    """Rectangle englobant l'encre réelle : traits de grille et texte.

    Le blanc laissé par le surdimensionnement n'est pas neutre : le vectorizer
    redimensionne l'image entière, donc chaque pixel vide consomme du budget de
    patches et rétrécit d'autant le texte utile.
    """
    rect = None
    for d in page.get_drawings():
        rect = d["rect"] if rect is None else rect | d["rect"]
    for w in page.get_text("words"):
        wr = pymupdf.Rect(w[:4])
        rect = wr if rect is None else rect | wr
    if rect is None:
        return None
    return pymupdf.Rect(rect.x0 - TRIM_PAD_PT, rect.y0 - TRIM_PAD_PT,
                        rect.x1 + TRIM_PAD_PT, rect.y1 + TRIM_PAD_PT) & page.rect


def _render_band(src: Path, sheet: Sheet, band: list[Tile], tmp: Path,
                 profile_dir: Path, dpi: int, tag: str) -> list[dict[str, Any]]:
    """Convertit une bande en une passe et découpe ses pages en tuiles.

    Retourne une liste VIDE si la pagination ne correspond pas au plan :
    l'appelant reprend alors tuile par tuile. On ne devine jamais quelle page
    porte quelle plage — livrer une image dont on ignore la plage est pire que
    payer N conversions.
    """
    extra = (sheet.width_px(sheet.min_col, sheet.min_col + sheet.freeze_cols - 1)
             if sheet.freeze_cols and band[0].col_start > sheet.min_col else 0.0)
    staged = tmp / f"stage_{tag}.xlsx"
    _stage(src, sheet, band, staged, extra)

    with pymupdf.open(to_pdf(staged, tmp, profile_dir)) as doc:
        if len(doc) != len(band):
            return []
        out = []
        for tile, page in zip(band, doc):
            pix = page.get_pixmap(dpi=dpi, clip=_clip(page))
            png = pix.tobytes("png")
            out.append({
                "index": tile.index,
                "range_a1": tile.range_a1,
                "row_start": tile.row_start, "row_end": tile.row_end,
                "col_start": tile.col_start, "col_end": tile.col_end,
                "width_px": pix.width, "height_px": pix.height,
                "sha256": hashlib.sha256(png).hexdigest(),
                "png_base64": base64.b64encode(png).decode("ascii"),
            })
        return out


def render_workbook(src: str | Path, *, dpi: int = DPI) -> dict[str, Any]:
    """Découpe tout le classeur et retourne le manifeste, images comprises.

    Chaque tuile porte sa plage A1 exacte ET son PNG en base64 : c'est ce lien
    qui permet à l'aval de rendre à l'utilisateur le morceau d'Excel
    correspondant à l'information citée.
    """
    src = Path(src)
    sheets = profile_workbook(src)
    out_sheets: list[dict[str, Any]] = []

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        profile_dir = seed_profile(tmp / "loprofile")

        for sheet in sheets:
            plan = plan_tiles(sheet)
            bands: dict[tuple[int, int], list[Tile]] = {}
            for t in plan:
                bands.setdefault((t.col_start, t.col_end), []).append(t)

            tiles: list[dict[str, Any]] = []
            for i, band in enumerate(bands.values()):
                got = _render_band(src, sheet, band, tmp, profile_dir, dpi,
                                   f"{len(out_sheets)}_{i}")
                if not got:
                    for t in band:
                        got += _render_band(src, sheet, [t], tmp, profile_dir,
                                            dpi, f"{len(out_sheets)}_{i}_{t.index}")
                tiles += got

            tiles.sort(key=lambda t: t["index"])
            out_sheets.append({
                "sheet": sheet.name,
                "range_a1": sheet.range_a1,
                "n_tiles": len(tiles),
                "tiles": tiles,
            })

    return {
        "source": src.name,
        "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "dpi": dpi,
        "data_uri_prefix": PNG_DATA_URI_PREFIX,
        "sheets": out_sheets,
    }
