"""
Ce qu'il faut savoir d'une feuille avant de la découper.

Tout passe par openpyxl plutôt que par un parseur OOXML maison : beaucoup moins
de code, au prix du chargement complet du classeur en mémoire. Arbitrage
assumé — ce module ne convient pas à un classeur de plusieurs centaines de Mo.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

__all__ = ["Sheet", "profile_workbook"]

DEFAULT_COL_WIDTH = 8.43        # caractères
DEFAULT_ROW_HEIGHT = 15.0       # points

# Mesuré sur LibreOffice 26 / Calibri 11 : px = 7.4 * largeur, sans padding.
# C'est une constante figée, pas une vérité : elle dépend de la police et de la
# version de LibreOffice. On peut se le permettre parce que la page de rendu est
# surdimensionnée puis rognée au contenu réel — une erreur de quelques pour cent
# sur la géométrie prédite ne se voit nulle part dans l'image finale.
PX_PER_CHAR = 7.4
PX_PER_POINT = 96 / 72

MAX_WIDTH_CHARS = 45.0          # plafond d'élargissement d'une colonne
WIDTH_SAMPLE_ROWS = 300         # lignes échantillonnées pour estimer les largeurs
EMU_PER_ROW = 9525 * 20         # 1 ligne par défaut ~ 20 px, pour les ancres


@dataclass
class Sheet:
    """Géométrie d'une feuille, en unités Excel. Les pixels se déduisent."""
    name: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    col_widths: dict[int, float]    # index -> largeur en caractères (0 = masquée)
    row_heights: dict[int, float]   # index -> hauteur en points (0 = masquée)
    header_rows: int                # lignes d'en-tête répétées sur chaque tuile
    freeze_cols: int
    row_breaks: list[int]
    anchors: list[tuple[int, int]]  # (ligne début, ligne fin) des graphiques/images

    def col_px(self, c: int) -> float:
        return self.col_widths.get(c, DEFAULT_COL_WIDTH) * PX_PER_CHAR

    def row_px(self, r: int) -> float:
        return self.row_heights.get(r, DEFAULT_ROW_HEIGHT) * PX_PER_POINT

    def width_px(self, c0: int, c1: int) -> float:
        return sum(self.col_px(c) for c in range(c0, c1 + 1))

    def height_px(self, r0: int, r1: int) -> float:
        return sum(self.row_px(r) for r in range(r0, r1 + 1))

    @property
    def range_a1(self) -> str:
        return (f"{get_column_letter(self.min_col)}{self.min_row}"
                f":{get_column_letter(self.max_col)}{self.max_row}")


def _display_len(value: object) -> float:
    """Longueur affichée d'une valeur, volontairement surestimée.

    Une colonne trop étroite n'affiche pas un texte tronqué : elle affiche
    `###`. C'est une perte TOTALE d'information dans le screenshot, bien pire
    qu'un libellé coupé. Or `str(32370.0)` fait 7 caractères là où Excel affiche
    `$32,370.00`, soit 10. On majore donc les nombres plutôt que de risquer la
    perte — une colonne trop large ne coûte que des pixels.
    """
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 5.0
    if isinstance(value, (datetime, date, time)):
        # str() rendrait "2014-01-01 00:00:00" (19) là où Excel affiche
        # "1/1/2014" (8) : on plafonne au format long usuel.
        return 12.0
    if isinstance(value, (int, float)):
        # Séparateurs de milliers et décimales, plus deux caractères pour un
        # éventuel symbole monétaire : c'est ce qu'Excel affichera au plus large.
        return len(f"{value:,.2f}") + 2.0
    return float(len(str(value)))


def _scan_cells(raw, val) -> tuple[tuple[int, int, int, int] | None, dict[int, float]]:
    """Bornes réelles et largeur nécessaire par colonne, en une passe.

    On lit les DEUX vues du classeur. La valeur en cache donne la largeur
    d'affichage ; la formule brute garantit qu'une colonne entièrement calculée
    ne passe pas pour vide — openpyxl rend None pour une formule sans valeur en
    cache, et cette colonne serait alors rognée hors de toutes les tuiles.

    `ws.max_row` ne suffit pas : il compte les cellules seulement formatées, ce
    qui produirait des tuiles vides à la fin de la feuille.
    """
    min_r = min_c = None
    max_r = max_c = 0
    needed: dict[int, float] = {}

    for i, (rrow, vrow) in enumerate(
            zip(raw.iter_rows(values_only=True),
                val.iter_rows(values_only=True)), start=1):
        for j, (rv, vv) in enumerate(zip(rrow, vrow), start=1):
            if rv is None and vv is None:
                continue
            if min_r is None:
                min_r = i
            min_c = j if min_c is None else min(min_c, j)
            max_r, max_c = i, max(max_c, j)
            if i - min_r < WIDTH_SAMPLE_ROWS:
                length = _display_len(vv if vv is not None else rv)
                if length > needed.get(j, 0.0):
                    needed[j] = length

    if min_r is None:
        return None, {}
    return (min_r, max_r, min_c, max_c), needed


def _column_widths(ws, min_col: int, max_col: int,
                   needed: dict[int, float]) -> dict[int, float]:
    """Largeur retenue par colonne : le maximum entre l'existante et le besoin.

    La largeur voulue par l'auteur est un plancher, jamais un plafond : on
    élargit pour ne rien perdre, on ne rétrécit jamais.
    """
    explicit: dict[int, float] = {}
    hidden: set[int] = set()
    for dim in ws.column_dimensions.values():
        lo = dim.min or column_index_from_string(dim.index or "A")
        hi = dim.max or lo
        for c in range(lo, min(hi, max_col) + 1):
            if dim.hidden:
                hidden.add(c)
            elif dim.width:
                explicit[c] = dim.width

    out: dict[int, float] = {}
    for c in range(min_col, max_col + 1):
        if c in hidden:
            out[c] = 0.0
            continue
        want = max(explicit.get(c, DEFAULT_COL_WIDTH),
                   min(needed.get(c, 0.0) + 1.5, MAX_WIDTH_CHARS))
        out[c] = round(want, 4)
    return out


def _row_heights(ws, min_row: int, max_row: int) -> dict[int, float]:
    """Hauteurs personnalisées. Les ignorer fausserait la hauteur des tuiles."""
    out: dict[int, float] = {}
    for r, dim in ws.row_dimensions.items():
        if not isinstance(r, int) or not (min_row <= r <= max_row):
            continue
        if dim.hidden:
            out[r] = 0.0
        elif dim.height:
            out[r] = dim.height
    return out


def _anchors(ws) -> list[tuple[int, int]]:
    """Lignes occupées par un graphique ou une image — zones incoupables.

    Couper un graphique en deux produit deux tuiles inexploitables : ni l'une ni
    l'autre ne porte l'information, et l'utilisateur reçoit une demi-image.
    """
    spans: list[tuple[int, int]] = []
    for obj in list(getattr(ws, "_charts", [])) + list(getattr(ws, "_images", [])):
        anchor = getattr(obj, "anchor", None)
        frm = getattr(anchor, "_from", None)
        if frm is None:
            continue
        r0 = frm.row + 1                      # openpyxl compte à partir de 0
        to = getattr(anchor, "to", None)
        if to is not None:
            r1 = to.row + 1
        else:
            # oneCellAnchor : pas de coin bas-droit, l'étendue est en EMU
            ext = getattr(anchor, "ext", None)
            cy = getattr(ext, "cy", 0) if ext is not None else 0
            r1 = r0 + max(1, int(cy / EMU_PER_ROW + 0.999)) - 1
        spans.append((r0, max(r0, r1)))
    return spans


def _header_rows(ws) -> tuple[int, int]:
    """-> (lignes d'en-tête à répéter, colonnes clés à répéter)

    Table déclarée d'abord, volets figés à défaut : une table Excel dit
    explicitement combien de lignes forment son en-tête.
    """
    header = 0
    for table in getattr(ws, "tables", {}).values():
        header = max(header, table.headerRowCount or 0)

    freeze_cols = 0
    pane = ws.freeze_panes
    if pane:
        cell = ws[pane] if isinstance(pane, str) else pane
        col, row = cell.column, cell.row
        freeze_cols = max(0, col - 1)
        if not header:
            header = max(0, row - 1)
    return header, freeze_cols


def profile_workbook(src: str | Path) -> list[Sheet]:
    """Profil géométrique de chaque feuille visible et non vide."""
    src = Path(src)
    wb_raw = openpyxl.load_workbook(src)                    # formules visibles
    wb_val = openpyxl.load_workbook(src, data_only=True)    # valeurs en cache

    sheets: list[Sheet] = []
    try:
        for name in wb_raw.sheetnames:
            ws = wb_raw[name]
            if ws.sheet_state != "visible":
                continue
            bounds, needed = _scan_cells(ws, wb_val[name])
            if bounds is None:
                continue                                     # feuille vide
            min_row, max_row, min_col, max_col = bounds
            header, freeze_cols = _header_rows(ws)
            sheets.append(Sheet(
                name=name,
                min_row=min_row, max_row=max_row,
                min_col=min_col, max_col=max_col,
                col_widths=_column_widths(ws, min_col, max_col, needed),
                row_heights=_row_heights(ws, min_row, max_row),
                header_rows=header,
                freeze_cols=freeze_cols,
                row_breaks=sorted({b.id for b in ws.row_breaks.brk if b.id}),
                anchors=_anchors(ws),
            ))
    finally:
        wb_raw.close()
        wb_val.close()
    return sheets
